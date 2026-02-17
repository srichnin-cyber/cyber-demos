#!/usr/bin/env python3
"""
Create sample Excel template files for testing Excel rendering examples.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import os

# Create templates directory if not exists
os.makedirs("src/main/resources/common-templates/templates", exist_ok=True)

def style_header(cell, text=""):
    """Apply header styling to a cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if text:
        cell.value = text
    return cell

def style_data(cell, value=""):
    """Apply data cell styling"""
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if value:
        cell.value = value
    return cell

# =============================================================================
# 1. Personal Form Template
# =============================================================================
print("Creating personal-form.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Personal Info"

# Headers
style_header(ws['A1'], "First Name")
style_header(ws['B1'], "Last Name")
style_header(ws['C1'], "Email")

# Data rows (leave empty for filling)
for row in range(2, 10):
    style_data(ws[f'A{row}'])
    style_data(ws[f'B{row}'])
    style_data(ws[f'C{row}'])

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30

wb.save("src/main/resources/common-templates/templates/personal-form.xlsx")
print("✅ personal-form.xlsx created")

# =============================================================================
# 2. Employee Roster Template
# =============================================================================
print("Creating employee-roster.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Employees"

# Headers
style_header(ws['A1'], "Employee ID")
style_header(ws['B1'], "First Name")
style_header(ws['C1'], "Last Name")
style_header(ws['D1'], "Email")
style_header(ws['E1'], "Department")

# Data rows (leave empty for filling)
for row in range(2, 52):
    for col in ['A', 'B', 'C', 'D', 'E']:
        style_data(ws[f'{col}{row}'])

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 20

wb.save("src/main/resources/common-templates/templates/employee-roster.xlsx")
print("✅ employee-roster.xlsx created")

# =============================================================================
# 3. Invoice Template
# =============================================================================
print("Creating invoice-template.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Invoice"

# Title
title_cell = ws['A1']
title_cell.value = "INVOICE"
title_cell.font = Font(bold=True, size=16)
ws.merge_cells('A1:D1')

# Header Info
style_header(ws['A3'], "Invoice #")
style_data(ws['B3'])

style_header(ws['C3'], "Invoice Date")
style_data(ws['D3'])

style_header(ws['A4'], "Due Date")
style_data(ws['B4'])

style_header(ws['C4'], "Customer")
style_data(ws['D4'])

# Line items header
style_header(ws['A6'], "Description")
style_header(ws['B6'], "Quantity")
style_header(ws['C6'], "Unit Price")
style_header(ws['D6'], "Total")

# Line items rows (leave empty for filling)
for row in range(7, 27):
    for col in ['A', 'B', 'C', 'D']:
        style_data(ws[f'{col}{row}'])

# Summary section
style_header(ws['A28'], "Subtotal")
style_data(ws['B28'])

style_header(ws['A29'], "Tax")
style_data(ws['B29'])

style_header(ws['A30'], "TOTAL")
total_cell = ws['B30']
total_cell.font = Font(bold=True, size=12)
total_cell.fill = PatternFill(start_color="E8F4EA", end_color="E8F4EA", fill_type="solid")

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

wb.save("src/main/resources/common-templates/templates/invoice-template.xlsx")
print("✅ invoice-template.xlsx created")

# =============================================================================
# 4. Generic Template (for range/row examples)
# =============================================================================
print("Creating template.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Headers for range examples
style_header(ws['A1'], "Item Names")
style_header(ws['B1'], "Item Prices")
style_header(ws['C1'], "Item Code")

# Pre-fill some placeholder data
for row in range(2, 7):
    style_data(ws[f'A{row}'])
    style_data(ws[f'B{row}'])
    style_data(ws[f'C{row}'])

# Row headers
style_header(ws['B10'], "Header 1")
style_header(ws['C10'], "Header 2")
style_header(ws['D10'], "Header 3")
style_header(ws['E10'], "Header 4")

# Matrix example (rows 13-15)
ws['A12'].value = "Matrix Example"
ws['A12'].font = Font(bold=True)

for row in range(13, 16):
    for col in ['A', 'B', 'C']:
        style_data(ws[f'{col}{row}'])

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

wb.save("src/main/resources/common-templates/templates/template.xlsx")
print("✅ template.xlsx created")

# =============================================================================
# 5. Employee Table Template (for table population example)
# =============================================================================
print("Creating employee-table.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Roster"

# Add some description at top
ws['A1'].value = "Employee Information"
ws['A1'].font = Font(bold=True, size=12)

# Headers at row 2
style_header(ws['A2'], "ID")
style_header(ws['B2'], "Name")
style_header(ws['C2'], "Department")
style_header(ws['D2'], "Salary")

# Empty data rows for population
for row in range(3, 25):
    for col in ['A', 'B', 'C', 'D']:
        style_data(ws[f'{col}{row}'])

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15

wb.save("src/main/resources/common-templates/templates/employee-table.xlsx")
print("✅ employee-table.xlsx created")

print("\n" + "="*60)
print("✅ All sample Excel templates created successfully!")
print("="*60)
print("\nTemplates created in:")
print("  src/main/resources/common-templates/templates/")
print("\nFiles created:")
print("  - personal-form.xlsx")
print("  - employee-roster.xlsx")
print("  - invoice-template.xlsx")
print("  - template.xlsx")
print("  - employee-table.xlsx")
print("\nReady to use with Excel rendering examples!")
