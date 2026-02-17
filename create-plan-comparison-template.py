#!/usr/bin/env python3
"""
Create a plan comparison Excel template.
This template is designed to be filled with a 2D matrix where:
- Row 1: Headers (Benefit label + plan names with spacing columns)
- Rows 2+: Benefit names + plan values
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_comparison_template():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Comparison"
    
    # Set column widths for readability
    sheet.column_dimensions['A'].width = 25  # Benefit name column
    sheet.column_dimensions['B'].width = 2   # First spacer
    sheet.column_dimensions['C'].width = 20  # Plan 1
    sheet.column_dimensions['D'].width = 2   # Spacer between Plan 1 and Plan 2
    sheet.column_dimensions['E'].width = 20  # Plan 2
    sheet.column_dimensions['F'].width = 2   # Spacer between Plan 2 and Plan 3
    sheet.column_dimensions['G'].width = 20  # Plan 3
    
    # Define header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define border
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Define data styling
    data_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Row 1: Header with benefit column and plan names
    headers = ["Benefit", "", "Plan A", "", "Plan B", "", "Plan C"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header_text
        if header_text and header_text != "":  # Don't style spacer columns
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        cell.border = border
    
    # Rows 2-6: Placeholder benefit rows (matrix will fill these)
    placeholder_benefits = [
        "Benefit 1",
        "Benefit 2", 
        "Benefit 3",
        "Benefit 4",
        "Benefit 5"
    ]
    
    for row_idx, benefit_name in enumerate(placeholder_benefits, 2):
        # Column A: Benefit name
        cell_a = sheet.cell(row=row_idx, column=1)
        cell_a.value = benefit_name
        cell_a.fill = data_fill
        cell_a.font = data_font
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.border = border
        
        # Columns B-G: Data cells (will be filled by matrix mapping)
        for col_idx in range(2, 8):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx % 2 == 0:  # Spacer columns (B, D, F)
                cell.value = ""
            else:  # Plan value columns (C, E, G)
                cell.value = ""  # Placeholder - will be filled
                cell.alignment = data_alignment
            cell.border = border
    
    # Save the template
    template_path = "src/main/resources/common-templates/templates/comparison-template.xlsx"
    workbook.save(template_path)
    print(f"✓ Created comparison template: {template_path}")

if __name__ == "__main__":
    create_comparison_template()
